"""Excel 写入模块

职责：
- 按 26 列格式生成 .xlsx 汇总表
- 包含施评文献信息 + 评论句 + 被评文献信息
"""

import logging
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from core.llm_analyzer import CommentRecord
from core.pdf_parser import PaperMetadata

logger = logging.getLogger(__name__)


# ── 作者名标准化工具 ──────────────────────────────────────────────

def _format_author_abbrev(full_name: str) -> str:
    """将英文作者全名转为 GB/T 7714 缩写格式

    规则：
    - 中文作者：原样返回
    - 英文 "First Middle Last" → "Last F M"
    - 英文 "Last, First Middle" → "Last F M"
    - 英文 "Last, F. M." → "Last F M"
    - 英文 "Last, F M" → "Last F M"
    - 已经是 "Last F" 格式 → 原样返回

    Examples:
        "Liumin Suo" → "Suo L"
        "Oleg Borodin" → "Borodin O"
        "Suo, L." → "Suo L"
        "Suo L" → "Suo L"
        "Teunis van Ree" → "van Ree T"
    """
    if not full_name or not full_name.strip():
        return full_name

    name = full_name.strip()

    # 中文作者直接返回
    if any('\u4e00' <= c <= '\u9fff' for c in name):
        return name

    # 常见姓氏前缀
    prefixes = {'van', 'von', 'de', 'del', 'della', 'der', 'den', 'di', 'la', 'le', 'el', 'al', 'bin', 'ibn'}

    if ',' in name:
        # 格式: "Last, First Middle" 或 "Last, F. M." 或 "Last, F.M."
        parts = name.split(',', 1)
        surname = parts[0].strip()
        given = parts[1].strip() if len(parts) > 1 else ""
        if given:
            # 提取所有名的首字母（处理 "J.R." → "JR"，"John" → "J"）
            given_clean = given.replace('.', ' ')
            given_parts = given_clean.split()
            initials = ''.join(p[0].upper() for p in given_parts if p)
            return f"{surname} {initials}" if initials else surname
        return surname
    else:
        # 格式: "First Middle Last" 或 "First Last"
        parts = name.split()
        if len(parts) == 1:
            return name

        # 检查是否已经是 "Last F" 或 "Last F M" 格式
        # 判断：最后一个部分是单字母或全大写单字母
        if all(len(p) == 1 and p.isupper() for p in parts[1:]):
            return name  # 已经是缩写格式

        # 找姓氏起始位置（处理 van der Berg 等复合姓氏）
        surname_start = len(parts) - 1
        for i in range(1, len(parts) - 1):
            if parts[i].lower() in prefixes:
                surname_start = i
                break

        surname = ' '.join(parts[surname_start:])
        given_parts = parts[:surname_start]
        initials = ''.join(p[0].upper() for p in given_parts if p)
        if initials:
            return f"{surname} {initials}"
        return surname


def _format_authors_abbrev(authors: list[str]) -> str:
    """将作者列表格式化为 GB/T 7714 缩写格式"""
    return ", ".join(_format_author_abbrev(a) for a in authors)


def _clean_text(text: str) -> str:
    """清理文本中的非法字符（openpyxl 不允许控制字符）"""
    if not text:
        return text
    # 移除所有 ASCII 控制字符（0x00-0x1F），除了换行(0x0A)和回车(0x0D)
    return re.sub(r'[\x00-\x09\x0b\x0c\x0e-\x1f]', '', text)


# 26 列表头
HEADERS = [
    "编号",
    "施评文献全部信息",
    "施评文献第一作者",
    "施评文献其他作者",
    "施评文章名",
    "施评期刊年卷期页码",
    "施评期刊",
    "施评年份",
    "施评卷期起止页码",
    "施评文献第一作者机构",
    "施评国家（第一作者）",
    "评论句",
    "标志词",
    "被评文献全部信息",
    "被评文献第一作者",
    "被评文献其他作者",
    "被评文章名",
    "被评期刊年卷期起止页码",
    "被评期刊全称",
    "被评年份",
    "被评卷期起止页码",
    "被评文献第一作者机构",
    "被评国家（第一作者）",
    "其他被评文献",
    "提供者",
    "备注",
]


def _format_citation(meta: PaperMetadata) -> str:
    """格式化施评文献的完整引用信息（GB/T 7714 格式）

    标准格式（参考示例）：
    作者列表（换行）标题.期刊, 年, 卷: 页码.
    """
    if meta.authors_cn:
        authors = ", ".join(meta.authors_cn)
    else:
        authors = ", ".join(meta.authors_en)
    title = meta.title_cn or meta.title_en
    journal = meta.journal_cn or meta.journal_en

    result = f"{authors}. {title}[J]. {journal}, {meta.year}"
    if meta.volume:
        result += f", {meta.volume}"
        if meta.issue:
            result += f"({meta.issue})"
    if meta.pages:
        result += f": {meta.pages}"
    result += "."
    return result


def _format_vol_issue_pages(meta: PaperMetadata) -> str:
    """格式化施评卷期起止页码（不含年份）

    标准格式：4: 1-34 或 4(2): 1-34
    """
    result = ""
    if meta.volume:
        result = meta.volume
        if meta.issue:
            result += f"({meta.issue})"
    if meta.pages:
        result += f": {meta.pages}"
    return result


def _format_journal_year_vol(meta: PaperMetadata) -> str:
    """格式化施评期刊年卷期页码（Col 6）

    标准格式：Electrochemical Energy Reviews, 2021, 4: 1-34.
    """
    journal = meta.journal_cn or meta.journal_en
    result = journal
    if meta.year:
        result += f", {meta.year}"
    if meta.volume:
        result += f", {meta.volume}"
        if meta.issue:
            result += f"({meta.issue})"
    if meta.pages:
        result += f": {meta.pages}"
    result += "."
    return result


def _format_evaluated_ref(record: CommentRecord) -> str:
    """格式化被评文献的完整引用（GB/T 7714 缩写格式）

    标准格式：Suo L, Borodin O, Gao T, ... "Title"[J]. Journal, Year, Vol(Issue): Pages.
    """
    ep = record.被评文献
    if ep.全部作者列表:
        authors = _format_authors_abbrev(ep.全部作者列表)
    else:
        authors = _format_author_abbrev(ep.第一作者)
    result = authors
    if ep.文章名:
        result += f". {ep.文章名}[J]."
    if ep.期刊名称:
        result += f" {ep.期刊名称},"
    if ep.年份:
        result += f" {ep.年份},"
    vol = ""
    if ep.卷:
        vol = ep.卷
        if ep.期:
            vol += f"({ep.期})"
    if vol:
        result += f" {vol}"
    if ep.起止页码:
        result += f": {ep.起止页码}"
    result += "."
    return result


def _format_evaluated_vol(record: CommentRecord) -> str:
    """格式化被评文献的期刊年卷期页码"""
    ep = record.被评文献
    result = ep.期刊名称 or ""
    if ep.年份:
        result += f", {ep.年份}"
    vol = ""
    if ep.卷:
        vol = ep.卷
        if ep.期:
            vol += f"({ep.期})"
    if vol:
        result += f", {vol}"
    if ep.起止页码:
        result += f": {ep.起止页码}"
    result += "."
    return result


def _format_evaluated_vol_pages(record: CommentRecord) -> str:
    """格式化被评文献的卷期起止页码"""
    ep = record.被评文献
    vol = ""
    if ep.卷:
        vol = ep.卷
        if ep.期:
            vol += f"({ep.期})"
    if ep.起止页码:
        vol += f": {ep.起止页码}"
    if vol:
        vol += "."
    return vol


def write_excel(
    output_path: str,
    records: list[CommentRecord],
    metadata: PaperMetadata,
    institution_results: list[dict] | None = None,
    provider: str = "",
) -> str:
    """生成 Excel 汇总表

    Args:
        output_path: 输出文件路径
        records: 评论句记录列表
        metadata: 施评文献元数据
        institution_results: 机构查询结果列表（与 records 一一对应）
        provider: 提供者信息

    Returns:
        输出文件路径
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # 样式定义
    header_font = Font(name="Times New Roman", bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    cell_font = Font(name="Times New Roman", size=11)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    wrap_alignment = Alignment(wrap_text=True, vertical="center")

    # 写入表头
    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = wrap_alignment

    # 预计算施评文献信息（每行相同）
    citation = _format_citation(metadata)
    journal_year_vol = _format_journal_year_vol(metadata)
    vol_issue_pages = _format_vol_issue_pages(metadata)

    # 写入数据行
    for i, record in enumerate(records):
        row = i + 2
        ep = record.被评文献

        # 机构查询结果
        inst = {}
        if institution_results and i < len(institution_results):
            inst = institution_results[i]

        # 被评作者机构和国家：优先用 LLM 返回的，其次用 CrossRef 查到的
        eval_institution = ep.第一作者机构 or inst.get("institution", "")
        eval_country = ep.第一作者国家 or inst.get("country", "")

        row_data = [
            i + 1,  # 编号
            citation,  # 施评文献全部信息
            metadata.first_author,  # 施评文献第一作者
            metadata.other_authors,  # 施评文献其他作者
            metadata.title_cn or metadata.title_en,  # 施评文章名
            journal_year_vol,  # 施评期刊年卷期页码
            metadata.journal_cn or metadata.journal_en,  # 施评期刊
            metadata.year,  # 施评年份
            vol_issue_pages,  # 施评卷期起止页码
            metadata.institution_cn or metadata.institution_en,  # 施评机构
            metadata.country,  # 施评国家
            record.评论句原文,  # 评论句
            record.标志词,  # 标志词
            _format_evaluated_ref(record),  # 被评文献全部信息
            _format_author_abbrev(ep.第一作者),  # 被评第一作者
            _format_authors_abbrev([a for a in ep.全部作者列表[1:]] if ep.全部作者列表 else [ep.其他作者] if ep.其他作者 else []),  # 被评其他作者
            ep.文章名,  # 被评文章名
            _format_evaluated_vol(record),  # 被评期刊年卷期页码
            ep.期刊名称,  # 被评期刊全称
            ep.年份,  # 被评年份
            _format_evaluated_vol_pages(record),  # 被评卷期起止页码
            eval_institution,  # 被评机构
            eval_country,  # 被评国家
            "",  # 其他被评文献
            provider,  # 提供者
            "",  # 备注
        ]

        for col, value in enumerate(row_data, 1):
            if isinstance(value, str):
                value = _clean_text(value)
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = cell_font
            cell.border = thin_border
            cell.alignment = wrap_alignment

    # 调整列宽
    col_widths = {
        1: 6, 2: 40, 3: 12, 4: 25, 5: 30,
        6: 30, 7: 15, 8: 8, 9: 18, 10: 25,
        11: 10, 12: 50, 13: 12, 14: 40, 15: 12,
        16: 25, 17: 30, 18: 30, 19: 15, 20: 8,
        21: 18, 22: 25, 23: 10, 24: 30, 25: 20, 26: 15,
    }
    for col, width in col_widths.items():
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    wb.save(output_path)
    logger.info(f"Excel 汇总表已保存: {output_path}")
    return output_path


def write_merged_excel(
    output_path: str,
    paper_results: list[dict],
) -> str:
    """将多篇论文的评论句合并到一个 Excel 文件

    Args:
        output_path: 输出文件路径
        paper_results: 每篇论文的处理结果列表，每个 dict 包含:
            - records: list[CommentRecord]
            - metadata: PaperMetadata
            - institution_results: list[dict] | None
            - provider: str

    Returns:
        输出文件路径
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "汇总"

    # 样式定义
    header_font = Font(name="Times New Roman", bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    cell_font = Font(name="Times New Roman", size=11)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    wrap_alignment = Alignment(wrap_text=True, vertical="center")

    # 写入表头
    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = wrap_alignment

    # 全局行号（跨论文连续编号）
    global_row = 2
    global_index = 1

    for paper in paper_results:
        records = paper["records"]
        metadata = paper["metadata"]
        institution_results = paper.get("institution_results")
        provider = paper.get("provider", "")

        if not records:
            continue

        # 预计算施评文献信息
        citation = _format_citation(metadata)
        journal_year_vol = _format_journal_year_vol(metadata)
        vol_issue_pages = _format_vol_issue_pages(metadata)

        for i, record in enumerate(records):
            ep = record.被评文献

            inst = {}
            if institution_results and i < len(institution_results):
                inst = institution_results[i]

            eval_institution = ep.第一作者机构 or inst.get("institution", "")
            eval_country = ep.第一作者国家 or inst.get("country", "")

            row_data = [
                global_index,
                citation,
                metadata.first_author,
                metadata.other_authors,
                metadata.title_cn or metadata.title_en,
                journal_year_vol,
                metadata.journal_cn or metadata.journal_en,
                metadata.year,
                vol_issue_pages,
                metadata.institution_cn or metadata.institution_en,
                metadata.country,
                record.评论句原文,
                record.标志词,
                _format_evaluated_ref(record),
                _format_author_abbrev(ep.第一作者),
                _format_authors_abbrev([a for a in ep.全部作者列表[1:]] if ep.全部作者列表 else [ep.其他作者] if ep.其他作者 else []),
                ep.文章名,
                _format_evaluated_vol(record),
                ep.期刊名称,
                ep.年份,
                _format_evaluated_vol_pages(record),
                eval_institution,
                eval_country,
                "",
                provider,
                "",
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=global_row, column=col, value=value)
                cell.font = cell_font
                cell.border = thin_border
                cell.alignment = wrap_alignment

            global_row += 1
            global_index += 1

    # 调整列宽
    col_widths = {
        1: 6, 2: 40, 3: 12, 4: 25, 5: 30,
        6: 30, 7: 15, 8: 8, 9: 18, 10: 25,
        11: 10, 12: 50, 13: 12, 14: 40, 15: 12,
        16: 25, 17: 30, 18: 30, 19: 15, 20: 8,
        21: 18, 22: 25, 23: 10, 24: 30, 25: 20, 26: 15,
    }
    for col, width in col_widths.items():
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    wb.save(output_path)
    total_records = global_index - 1
    logger.info(f"合并 Excel 已保存: {output_path}（共 {total_records} 条记录）")
    return output_path
